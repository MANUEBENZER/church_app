from django.contrib import messages
from django.utils import timezone
from datetime import date, timedelta
from django.shortcuts import render, redirect
from django.utils.timezone import now
from django.db.models import Count,  Q, Sum
from django.http import HttpResponse, JsonResponse
from django.template.loader import get_template
from lxml.objectify import annotate
from xhtml2pdf import pisa
import openpyxl
from django.http import HttpResponse
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404
from .models  import Attendance, Member, DailyChurchReport, Promise, Income, Expense, Announcement
from django.views.generic import TemplateView
from django.contrib.auth.mixins import UserPassesTestMixin
from .utils import is_admin, is_secretary
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.utils.decorators import method_decorator
from django.core.paginator import Paginator
from reportlab.pdfgen import canvas
from django import forms

from django.contrib.auth.decorators import login_required, user_passes_test





@login_required
def post_login_redirect(request):
    user = request.user

    # Django Admin users
    if user.is_superuser or user.is_staff:
        return redirect("/admin/")

    if user.groups.filter(name="Secretary").exists():
        return redirect("members:secretary_dashboard")

    if user.groups.filter(name="Pastor").exists():
        return redirect("members:pastor_dashboard")

    # fallback
    return redirect("members:home")


class AttendanceDashboardView(TemplateView):
    template_name = 'attendance/secretary_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        qs = Attendance.objects.all()

        # Filters
        start = self.request.GET.get('start')
        end = self.request.GET.get('end')
        service = self.request.GET.get('service')

        if start:
            qs = qs.filter(date__gte=start)
        if end:
            qs = qs.filter(date__lte=end)
        if service:
            qs = qs.filter(service=service)

        context['total'] = qs.count()
        context['present'] = qs.filter(present=True).count()
        context['absent'] = qs.filter(present=False).count()
        context['services'] = Attendance.objects.values_list('service', flat=True).distinct()

        return context


class PastorDashboardView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = "pastor/secretary_dashboard.html"

    def test_func(self):
        return self.request.user.groups.filter(name="Pastor").exists()


from django.db.models.functions import TruncWeek, TruncMonth

weekly = (
    Attendance.objects
    .annotate(week=TruncWeek('date'))
    .values('week')
    .annotate(total=Count('id'))
)

monthly = (
    Attendance.objects
    .annotate(month=TruncMonth('date'))
    .values('month')
    .annotate(total=Count('id'))
)


@login_required
def announcements_dashboard(request):
    today = now().date()

    report, created = DailyChurchReport.objects.get_or_create(
        date=today,
        defaults={
            'service_type': 'Sunday',
            'announcements': '',
            'total_attendance': 0,
            'offering_amount': 0,
            'created_by': request.user
        }
    )

    # 🔹 AUTO-PULL attendance summary
    attendance_summary = Attendance.objects.filter(
        date=today,
        service_type=report.service_type
    ).aggregate(
        total_present=Count('id', filter=Q(present=True)),
        total_absent=Count('id', filter=Q(present=False)),
        total_members=Count('id')
    )

    # 🔹 Auto-fill attendance
    report.total_attendance = attendance_summary['total_present'] or 0
    report.save(update_fields=['total_attendance'])

    if request.method == "POST":
        # 🔹 Update report fields
        report.service_type = request.POST.get("service_type")
        report.announcements = request.POST.get("announcements")
        report.offering_amount = request.POST.get("offering_amount") or 0
        report.promises_amount = request.POST.get("promises_amount") or 0
        report.special_notes = request.POST.get("special_notes")
        report.created_by = request.user
        report.save()

        # 🔹 Get promise fields safely
        member_name = request.POST.get("promise_member")
        promise_type = request.POST.get("promise_type")
        amount = request.POST.get("promise_amount")
        item = request.POST.get("promise_item")
        due_date = request.POST.get("promise_due_date")

        # 🔹 Save promise only if filled
        if member_name and promise_type:
            Promise.objects.create(
                report=report,
                member_name=member_name,
                promise_type=promise_type,
                amount=amount if promise_type == "money" else None,
                item_description=item if promise_type == "item" else "",
                due_date=due_date or None
            )

        return redirect("members:announcements_dashboard")
    promises = Promise.objects.filter(report=report)
    promise_filter = request.GET.get('promise_filter', 'all')
    if promise_filter == 'overdue':
        promises = promises.filter(due_date__lt=today, fulfilled=False)
    elif promise_filter == 'fulfilled':
        promises = promises.filter(fulfilled=True)
    elif promise_filter == 'unfulfilled':
        promises = promises.filter(fulfilled=False)

    return render(request, "announcements/announcements_dashboard.html", {
        "report": report,
        "today": today,
        "attendance_summary": attendance_summary,
        "promises": promises,
        "request": request,
    })

@login_required
def add_announcement(request): 
    if request.method == "POST":
        title = request.POST.get("title")
        message = request.POST.get("message")

        if title and message:
            Announcement.objects.create(
                title=title,
                message=message,
                created_by=request.user
            )
            return redirect("members:announcement_dashboard")

    return render(request, "announcements/add_announcement.html")

@login_required
@user_passes_test(is_secretary)
def announcement_edit(request, pk):
    announcement = get_object_or_404(Announcement, pk=pk)

    if request.method == "POST":
        announcement.title = request.POST.get("title")
        announcement.message = request.POST.get("message")
        announcement.save()
        return redirect("members:announcement_dashboard")

    return render(request, "announcements/announcement_edit.html", {
        "announcement": announcement
    })


@login_required
@user_passes_test(is_secretary)
def announcement_delete(request, pk):
    announcement = get_object_or_404(Announcement, pk=pk)

    if request.method == "POST":
        announcement.delete()
        return redirect("members:announcement_dashboard")

    return redirect("members:announcement_dashboard")
@login_required
@user_passes_test(is_secretary)
def announcement_slide_show(request):
    today = date.today()

    announcements = Announcement.objects.filter(
        date=today,
        is_active=True
    ).order_by("created_at")

    return render(request, "announcements/display.html", {
        "announcements": announcements,
        "today": today,
    })



def announcement_display(request):
    today = date.today()

    announcements = Announcement.objects.filter(
        date=today,
        is_active=True
    ).order_by("created_at")

    return render(request, "announcements/display.html", {
        "announcements": announcements,
        "today": today,
    })





@login_required
def export_announcements_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="announcements.pdf"'

    p = canvas.Canvas(response)
    y = 800

    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "Church Announcements")
    y -= 30

    p.setFont("Helvetica", 10)
    for ann in Announcement.objects.all().order_by("-created_at"):
        if y < 50:
            p.showPage()
            y = 800

        p.drawString(50, y, f"{ann.title} ({ann.created_at.strftime('%d %b %Y')})")
        y -= 15
        p.drawString(60, y, ann.message[:100])
        y -= 25

    p.showPage()
    p.save()
    return response













def is_admin(user):
    return user.groups.filter(name='Admin').exists()

@user_passes_test(is_admin)
def delete_attendance(request, id):
    ...



def attendance_chart_data(request):
    qs = Attendance.objects.all()

    if request.GET.get('service'):
        qs = qs.filter(service=request.GET['service'])

    data = (
        qs.values('date')
        .annotate(present=Count('id', filter=Q(present=True)))
        .order_by('date')
    )

    return JsonResponse({
        'labels': [d['date'] for d in data],
        'present': [d['present'] for d in data]
    })

class AdminRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return is_admin(self.request.user)


@require_POST
def ajax_toggle_attendance(request):
    attendance_id = request.POST.get('attendance_id')
    attendance = get_object_or_404(Attendance, id=attendance_id)

    attendance.present = not attendance.present
    attendance.save()

    return JsonResponse({
        'status': 'present' if attendance.present else 'absent'
    })
# -----------------------------
# Add Attendance (AJAX-friendly)
# -----------------------------
@login_required
def add_attendance(request):
    today = date.today()

    # --- GET PARAMETERS ---
    service_type = request.GET.get("service", "Sunday")
    search_query = request.GET.get("q", "").strip()
    selected_date_str = request.GET.get("date")
    selected_date = date.fromisoformat(selected_date_str) if selected_date_str else today

    is_locked = selected_date < today

    # --- GET MEMBERS ---
    members = Member.objects.all().order_by("full_name").prefetch_related("attendance_set")

    # Filter active members if needed
    members = members.filter(is_active=True)

    # Apply search filter if any
    if search_query:
        members = members.filter(
            Q(full_name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(membership_id__icontains=search_query)
        )

    members = list(members)  # Convert to list for iteration

    # --- ENSURE ATTENDANCE RECORDS EXIST ---
    existing_ids = set(
        Attendance.objects.filter(
            date=selected_date,
            service_type=service_type
        ).values_list("member_id", flat=True)
    )

    new_attendance = [
        Attendance(
            member=member,
            date=selected_date,
            service_type=service_type,
            present=False
        )
        for member in members
        if member.id not in existing_ids
    ]

    if new_attendance:
        Attendance.objects.bulk_create(new_attendance)

    # --- GET PRESENT MEMBER IDS ---
    present_member_ids = list(
        Attendance.objects.filter(
            date=selected_date,
            service_type=service_type,
            present=True
        ).values_list("member_id", flat=True)
    )

    # --- CALCULATE ATTENDANCE PERCENTAGE ---
    for member in members:
        attendance_records = member.attendance_set.all()
        total = attendance_records.count()
        present = attendance_records.filter(present=True).count()
        member.attendance_percentage = round((present / total) * 100, 1) if total else 0

    # --- SUMMARY ---
    total_members = len(members)
    total_present = len(present_member_ids)
    total_absent = total_members - total_present

    attendance_summary = {
        "total_members": total_members,
        "total_present": total_present,
        "total_absent": total_absent,
    }

    return render(request, "attendance/add_attendance.html", {
        "members": members,
        "service_type": service_type,
        "today": today,
        "selected_date": selected_date,
        "is_locked": is_locked,
        "present_member_ids": present_member_ids,
        "search_query": search_query,
        "attendance_summary": attendance_summary,
        "report": {"total_attendance": total_present},
    })

# -----------------------------
# AJAX Attendance Marking
# -----------------------------
@require_POST
def ajax_mark_attendance(request):
    member_id = request.POST.get("member_id")
    service_type = request.POST.get("service_type")
    attendance_date = request.POST.get("date")
    present = request.POST.get("present") == "true"

    # Basic validation
    if not all([member_id, service_type, attendance_date]):
        return JsonResponse(
            {"status": "error", "message": "Missing required fields"},
            status=400
        )


    attendance, created = Attendance.objects.get_or_create(
        member_id=member_id,
        service_type=service_type,
        date=attendance_date,
        defaults={"present": present},
    )

    if not created:
        attendance.present = present
        attendance.save(update_fields=["present"])

    return JsonResponse({
        "status": "ok",
        "created": created,
        "present": attendance.present,
    })
# -----------------------------
# Attendance List
# -----------------------------
#def attendance_list(request):
#    members = Member.objects.annotate(
#        total_attendance=Count('attendance'),
#        present_count=Count(
#            'attendance',
#            filter=Q(attendance__present=True)
#        )
#    )
#
#    return render(request, 'attendance/attendance_list.html', {
#        'members': members,
#    })
#def attendance_list(request):
 #   user = request.user
#
  #  if user.groups.filter(name="Secretary").exists():
#        base_template = "secretary/base.html"
#    elif user.groups.filter(name="Pastor").exists():
#        base_template = "pastor/base.html"
 #   elif user.is_superuser:
 #       base_template = "admin/base_site.html"
#    else:
#        base_template = "base.html"
#
 #   context = {
 #       "base_template": base_template,
 #       "members": members,
 #       "services": services,
 #   }
#
 #   return render(request, "attendance_list.html", context)



@login_required
def attendance_list(request):
    user = request.user

    if user.groups.filter(name="Secretary").exists():
        base_template = "secretary/base.html"
    elif user.groups.filter(name="Pastor").exists():
        base_template = "pastor/base.html"
    elif user.is_superuser:
        base_template = "admin/base_site.html"
    else:
        base_template = "base.html"

    members = Member.objects.annotate(
        total_attendance=Count('attendance'),
        present_count=Count(
            'attendance',
            filter=Q(attendance__present=True)
        )
    )

    # 🔹 Extract services from Attendance instead
    services = Attendance.objects.values_list(
        "service_type", flat=True
    ).distinct()

    context = {
        "base_template": base_template,
        "members": members,
        "services": services,
    }

    return render(request, "attendance/attendance_list.html", context)



# -----------------------------
# Dashboard
# -----------------------------
@login_required
def secretary_dashboard(request):
    today = now().date()
    # Last 7 days
    last_7_days = [today - timedelta(days=i) for i in range(6, -1, -1)]

    attendance_data = []
    for day in last_7_days:
        count = Attendance.objects.filter(date=day, present=True).count()
        attendance_data.append({'date': day.strftime('%Y-%m-%d'), 'count': count})

    total_members = Member.objects.count()

    new_members = Member.objects.filter(
        date_joined__year=today.year,
        date_joined__month=today.month
    ).count()


    present_today = Attendance.objects.filter(date=today, present=True).count()
    
    new_members_list = Member.objects.filter(
        date_joined__year=today.year,
        date_joined__month=today.month
    ).order_by("-date_joined")


    return render(request, 'secretary/secretary_dashboard.html', {
        'attendance_data': attendance_data,
        'total_members': total_members,
        'present_today': present_today,
        'new_members': new_members,
        'new_members_list': new_members_list,
    })












def export_members_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=church_members.pdf"

    p = canvas.Canvas(response)
    y = 800

    p.setFont("Helvetica-Bold", 14)
    p.drawString(200, y, "Church Members List")
    y -= 40

    p.setFont("Helvetica", 10)

    for m in Member.objects.all():
        p.drawString(
            50, y,
            f"{m.membership_id} | {m.full_name} | {m.department} | {m.phone}"
        )
        y -= 18
        if y < 50:
            p.showPage()
            y = 800

    p.save()
    return response




def export_members_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "Membership ID", "Full Name", "Gender", "Department",
        "Position", "Phone", "Email", "Status"
    ])

    for m in Member.objects.all():
        ws.append([
            m.membership_id,
            m.full_name,
            m.gender,
            m.department,
            m.position,
            m.phone,
            m.email,
            "Active" if m.is_active else "Inactive"
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=church_members.xlsx"
    wb.save(response)
    return response


@login_required
@user_passes_test(is_secretary)
def member_create(request):
    if request.method == "POST":
        form = MemberForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect("members:member_list")
    else:
        form = MemberForm()
    return render(request, "members/create_member_form.html", {"form": form, "action": "Add"})

@login_required
@user_passes_test(is_secretary)
def member_edit(request, pk):
    member = get_object_or_404(Member, pk=pk)

    if request.method == "POST":
        form = MemberForm(request.POST, request.FILES, instance=member)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.updated_by = request.user
            obj.save()
            return redirect("members:member_detail", pk=member.pk)
    else:
        form = MemberForm(instance=member)

    return render(request, "members/member_edit_detail.html", {
        "form": form,
        "member": member
    })

@login_required
#@user_passes_test(is_superuser)
def member_delete(request, pk):
    member = get_object_or_404(Member, pk=pk)
    if request.method == "POST":
        member.delete()
        messages.success(request, "Member deleted successfully.")
        return redirect("members:member_list")
    # Optional: confirm page
    return render(request, "members/member_confirm_delete.html", {"member": member})





# -----------------------------
# Export Attendance to Excel
# -----------------------------
def export_attendance_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Attendance + {{Service.name}}+' ' +{{Date}}"

    # Header
    ws.append(["Member", "Service", "Date", "Present"])

    # Data
    for att in Attendance.objects.select_related('member').all():
        ws.append([
            f"{att.member.full_name} ",
            att.service_type,
            att.date.strftime('%Y-%m-%d'),
            "Yes" if att.present else "No"
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=attendance.xlsx'
    wb.save(response)
    return response


# -----------------------------
# Export Attendance to PDF
# -----------------------------
def export_attendance_pdf(request):
    template_path = 'attendance_pdf.html'
    attendances = Attendance.objects.select_related('member').all()
    context = {'attendances': attendances}
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="attendance.pdf"'
    template = get_template(template_path)
    html = template.render(context)
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response



class SecretaryDashboardView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    member = Member.objects.count()
    template_name = "secretary/secretary_dashboard.html, { member: 'member'}"

    def test_func(self):
        user = self.request.user
        return user.groups.filter(name="Secretary").exists()


def member_list(request):
    q = request.GET.get("q", "")
    members = Member.objects.filter(full_name__icontains=q) if q else Member.objects.all()

    paginator = Paginator(members, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "members/member_list.html", {
        "members": members,
        "page_obj": page_obj
    })


@method_decorator(login_required, name="dispatch")
class SecretaryDashboardView(TemplateView):
    template_name = "secretary/secretary_dashboard.html"


@login_required
def announcement_list(request):
    return render(request, "members/secretary/announcement_list.html")


@login_required
def finance_dashboard(request):
    return render(request, "members/secretary/finance_dashboard.html")


@login_required
def reports(request):
    return render(request, "members/secretary/reports.html")

@login_required
def secretary_settings(request):
    return render(request, "members/secretary/secretary_settings.html")

@login_required
@user_passes_test(is_secretary)
def member_detail(request, pk):
    member = get_object_or_404(Member, pk=pk)

    attendance_summary = (
        Attendance.objects
        .filter(member=member, present=True)
        .values("service_type")
        .annotate(total=Count("id"))
    )

    total_attendance = Attendance.objects.filter(
        member=member, present=True
    ).count()

    return render(request, "members/member_detail.html", {
        "member": member,
        "attendance_count": total_attendance,
        "attendance_summary": attendance_summary
    })

class MemberForm(forms.ModelForm):
    class Meta:
        model = Member
        fields = "__all__"
        widgets = {
            "date_of_birth": forms.DateInput(attrs={"type": "date"}),
        }










# -----------------------------
# Home Redirect
# -----------------------------
def home(request):
    return render(request, "home.html")











@login_required
def promises_dashboard(request):
    q = request.GET.get("q")
    today = timezone.now().date()

    promises = Promise.objects.all().order_by("-due_date")

    if q:
        promises = promises.filter(
            member_name__icontains=q
        )

    total_money = promises.filter(
        promise_type="money"
    ).aggregate(total=Sum("amount"))["total"] or 0

    total_items = promises.filter(
        promise_type="item"
    ).count()
    
    service_type = request.GET.get("service", "Sunday")
    search_query = request.GET.get("q", "").strip()
    selected_date_str = request.GET.get("date")
    selected_date = date.fromisoformat(selected_date_str) if selected_date_str else today

    members = Member.objects.all().order_by("full_name").prefetch_related("attendance_set")
    present_member_ids = list(
        Attendance.objects.filter(
            date=selected_date,
            service_type=service_type,
            present=True
        ).values_list("member_id", flat=True)
    )

    for member in members:
        attendance_records = member.attendance_set.all()
        total = attendance_records.count()
        present = attendance_records.filter(present=True).count()
        member.attendance_percentage = round((present / total) * 100, 1) if total else 0
    total_members = len(members)
    total_present = len(present_member_ids)
    total_absent = total_members - total_present

    attendance_summary = {
        "total_members": total_members,
        "total_present": total_present,
        "total_absent": total_absent,
    }
    return render(request, "announcements/promises_dashboard.html", {
        "promises": promises,
        "today": today,
        "total_money": total_money,
        "total_items": total_items,
        "attendance_summary": attendance_summary,
        "present_member_ids": present_member_ids,
    })

@login_required
@user_passes_test(is_secretary)
def add_promise(request):
    # Get today's report (or adjust if you select report differently)
    report = DailyChurchReport.objects.filter(date=timezone.now().date()).first()

    if request.method == "POST":
        member_name = request.POST.get("member_name")
        promise_type = request.POST.get("promise_type")
        amount = request.POST.get("amount")
        item_description = request.POST.get("item")
        due_date = request.POST.get("promise_due_date")

        Promise.objects.create(
            report=report,
            member_name=member_name,
            promise_type=promise_type,
            amount=amount if promise_type == "money" else None,
            item_description=item_description if promise_type == "item" else "",
            due_date=due_date or None,
        )

        return redirect("members:promises_dashboard")

    return render(request, "announcements/add_promise.html")





@login_required
def edit_promise(request, promise_id):
    promise = get_object_or_404(Promise, id=promise_id)
    if request.method == 'POST':
        promise.member_name = request.POST.get('member_name', promise.member_name)
        promise.promise_type = request.POST.get('promise_type', promise.promise_type)
        promise.amount = request.POST.get('amount', promise.amount)
        promise.item_description = request.POST.get('item', promise.item_description)
        promise.due_date = request.POST.get('promise_due_date', promise.due_date)
        promise.fulfilled = 'fulfilled' in request.POST
        promise.save()
        return redirect('members:promises_dashboard')
    return render(request, 'announcements/edit_promise.html', {'promise': promise})

@login_required
def delete_promise(request, promise_id):
    promise = get_object_or_404(Promise, id=promise_id)
    if request.method == 'POST':
        promise.delete()
        return redirect('members:promises_dashboard')
    return render(request, 'announcements/delete_promise.html', {'promise': promise})



@login_required
def export_promises_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="promises.pdf"'

    p = canvas.Canvas(response)
    y = 800

    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "Church Promises")
    y -= 30

    p.setFont("Helvetica", 10)
    for pr in Promise.objects.all():
        if y < 50:
            p.showPage()
            y = 800

        p.drawString(
            50, y,
            f"{pr.member_name} - {pr.promise_type.upper()} - {pr.amount or pr.item_description}"
        )
        y -= 20

    p.save()
    return response




@login_required
def export_promises_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Promises"

    ws.append([
        "Member", "Type", "Amount", "Item", "Due Date", "Status"
    ])

    for p in Promise.objects.all():
        ws.append([
            p.member_name,
            p.promise_type,
            p.amount,
            p.item_description,
            p.due_date,
            "Fulfilled" if p.fulfilled else "Pending"
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=promises.xlsx'
    wb.save(response)
    return response




















@login_required
def finance_dashboard(request):
    today = date.today()
    current_year = today.year
    current_month = today.month

    # =========================
    # INCOME TOTALS
    # =========================
    total_income = Income.objects.aggregate(
        total=Sum("amount")
    )["total"] or 0

    monthly_income = Income.objects.filter(
        date__year=current_year,
        date__month=current_month
    ).aggregate(total=Sum("amount"))["total"] or 0

    income_by_type = (
        Income.objects
        .values("income_type")
        .annotate(total=Sum("amount"))
        .order_by("-total")
    )

    # =========================
    # EXPENSE TOTALS
    # =========================
    total_expense = Expense.objects.aggregate(
        total=Sum("amount")
    )["total"] or 0

    monthly_expense = Expense.objects.filter(
        date__year=current_year,
        date__month=current_month
    ).aggregate(total=Sum("amount"))["total"] or 0

    expense_by_category = (
        Expense.objects
        .values("category")
        .annotate(total=Sum("amount"))
        .order_by("-total")
    )

    # =========================
    # NET BALANCE
    # =========================
    net_balance = total_income - total_expense

    # =========================
    # RECENT TRANSACTIONS
    # =========================
    recent_incomes = Income.objects.order_by("-date")[:5]
    recent_expenses = Expense.objects.order_by("-date")[:5]

    return render(request, "finance/finance_dashboard.html", {
        "total_income": total_income,
        "total_expense": total_expense,
        "monthly_income": monthly_income,
        "monthly_expense": monthly_expense,
        "net_balance": net_balance,
        "income_by_type": income_by_type,
        "expense_by_category": expense_by_category,
        "recent_incomes": recent_incomes,
        "recent_expenses": recent_expenses,
        "current_month": today.strftime("%B"),
        "current_year": current_year,
    })
from django.utils.timezone import localtime

@login_required
def announcement_dashboard(request):
    today = date.today()

    if request.method == "POST":
        title = request.POST.get("title")
        message = request.POST.get("message")

        if title and message:
            Announcement.objects.create(
                title=title,
                message=message,
                created_by=request.user
            )
            return redirect("announcement_dashboard")
    def get_live_announcements():
        now = localtime()
        return Announcement.objects.filter(
            created_at=date.today(),
            start_time__lte=now.time(),
            end_time__gte=now.time(),
            is_active=True
        ).order_by("created_at")
    announcements = Announcement.objects.filter(
        date=today,
        is_active=True
    ).order_by("-created_at")
   
    return render(request, "announcements/announcement_dashboard.html", {
        "announcements": announcements,
        "today": today,
    })
#    
    






@login_required
def finance_home(request):
    return render(request, 'finance/finance_home.html') 







@login_required
def tithes_dashboard(request):
    return render(request, 'finance/tithes_dashboard.html')



from django.db.models import Sum, Max
from datetime import date

def tithes_dashboard(request):
    tithes = Income.objects.filter(income_type='tithe')

    date_query = request.GET.get('date')
    if date_query:
        tithes = tithes.filter(date=date_query)

    now_date = now()

    context = {
        'tithes': tithes,
        'current_month': now_date.strftime('%B'),
        'current_year': now_date.year,
        'monthly_total': tithes.filter(
            date__month=now_date.month,
            date__year=now_date.year
        ).aggregate(total=Sum('amount'))['total'] or 0,
        'total_records': tithes.count(),
        'highest_tithe': tithes.aggregate(max=Max('amount'))['max'] or 0,
    }

    return render(request, 'finance/tithes_dashboard.html', context)



from .forms import TitheForm

def add_tithe(request):
    if request.method == 'POST':
        form = TitheForm(request.POST)
        if form.is_valid():
            tithe = form.save(commit=False)
            tithe.income_type = 'tithe'          # force tithe
            tithe.recorded_by = request.user     # who recorded it
            tithe.save()

            messages.success(request, "Tithe recorded successfully!")
            return redirect('tithes_dashboard')
    else:
        form = TitheForm(initial={'date': now().date()})

    return render(request, 'finance/add_tithe.html', {'form': form})






# ================== Form ==================
class IncomeForm(forms.ModelForm):
    class Meta:
        model = Income
        fields = ['amount', 'income_type', 'date', 'description']
        widgets = {
            'date': forms.DateInput(attrs={'type': 'date'}),
            'description': forms.Textarea(attrs={'rows': 2}),
        }

# ================== Edit View ==================
def edit_tithe(request, pk):
    tithe = get_object_or_404(Income, pk=pk, income_type='tithe')
    
    if request.method == 'POST':
        form = IncomeForm(request.POST, instance=tithe)
        if form.is_valid():
            form.save()
            messages.success(request, "Tithe record updated successfully!")
            return redirect('tithes_records')
    else:
        form = IncomeForm(instance=tithe)
    
    context = {
        'form': form,
        'tithe': tithe
    }
    return render(request, 'edit_tithe.html', context)


def delete_tithe(request, pk):
    tithe = get_object_or_404(Income, pk=pk, income_type='tithe')
    
    if request.method == 'POST':
        tithe.delete()
        messages.success(request, "Tithe record deleted successfully!")
        return redirect('tithes_records')
    
    # If you want a confirmation page, render it
    return render(request, 'confirm_delete_tithe.html', {'tithe': tithe})


def tithes_pdf(request):
    # Filter tithe records
    tithes = Income.objects.filter(income_type='tithe')

    # Context similar to tithes_records view
    current_month = now().strftime('%B')
    current_year = now().year
    monthly_total = tithes.filter(date__month=now().month).aggregate(Sum('amount'))['amount__sum'] or 0
    total_records = tithes.count()
    highest_tithe = tithes.aggregate(Max('amount'))['amount__max'] or 0

    context = {
        'tithes': tithes,
        'current_month': current_month,
        'current_year': current_year,
        'monthly_total': monthly_total,
        'total_records': total_records,
        'highest_tithe': highest_tithe,
    }

    # Load template
    template_path = 'tithes_pdf.html'  # you will create a minimal PDF template
    template = get_template(template_path)
    html = template.render(context)

    # Create PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename="tithes_records.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF <pre>' + html + '</pre>')
    return response













@login_required
def dashboard(request):
    return render(request, 'secretary_dashboard.html')
